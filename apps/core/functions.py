import datetime
import os
from time import time
import requests
import json
from django.conf import settings
from openpyxl import Workbook
from django.http import JsonResponse
from django.db.models import QuerySet
from apps.account.models import GrantRequest, User, GrantRecord
from apps.lab.models import Request
from apps.order.models import PaymentRecord
import pandas as pd
import jdatetime
from django.utils import timezone
from django.contrib.auth import get_user_model

User = get_user_model()

def process_excel_and_create_grant_records(file):
    df = pd.read_excel(file)

    created_records = []

    for index, row in df.iterrows():
        national_id = row['receiver']
        try:
            receiver = User.objects.get(national_id=national_id)
        except User.DoesNotExist:
            continue
        shamsi_date = row['expiration_date']
        if pd.notna(shamsi_date):
            expiration_date = jdatetime.date(*map(int, shamsi_date.split('/'))).togregorian()
        else:
            expiration_date = None

        grant_record = GrantRecord.objects.create(
            title=row['title'],
            receiver=receiver,
            amount=row['amount'],
            expiration_date=expiration_date,
            created_at=timezone.now()
        )
        created_records.append(grant_record)

    return created_records


def labsnet(nid, type, service):
    import requests
    import json
    # data = {"user_name": "sharif_uni", "password": "sharif_uni", "national_code": 4723686509, "type": 1,
    #         "services": "[15737, 15737]"}

    sdata = '{"user_name": "sharif_uni", "password": "sharif_uni", "national_code": "4723686509", "type": "1", "services": [115901, 15737]}'
    jsdata = json.loads(sdata)
    response = requests.post('https://labsnet.ir/api/credit_list', data=jsdata, verify=False)
    response.text
    # response_text = '{"customer_name":"\\u0645\\u062d\\u0633\\u0646 \\u0686\\u06af\\u0646\\u06cc","credits":[]}'

    data = {"user_name": "sharif_uni", "password": "sharif_uni", "national_code": nid, "type": type, "services": list(service)}
    jdata = json.loads(data)
    response = requests.post('https://labsnet.ir/api/credit_list', data=jdata, verify=False)
    response_text = response.text
    return JsonResponse(json.loads(response_text))


def export_excel(queryset):
    if isinstance(queryset, QuerySet) and queryset.model == Request:
        row_list, columns, sheet_title = request()
    if isinstance(queryset, QuerySet) and queryset.model == GrantRequest:
        row_list, columns, sheet_title = grant_request()
    if isinstance(queryset, QuerySet) and queryset.model == User:
        row_list, columns, sheet_title = user()
    if isinstance(queryset, QuerySet) and queryset.model == PaymentRecord:
        row_list, columns, sheet_title = payment_record()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title

    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for obj in queryset.all():
        row_num += 1
        row = row_list(obj)

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    p2_name = os.path.join('exports', sheet_title, f'{sheet_title}-{datetime.datetime.now().strftime("%y%m%d-%H:%M:%S")}.xlsx')
    file_path = os.path.join(settings.MEDIA_ROOT, p2_name)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    workbook.save(file_path)


    file_url = os.path.join(settings.MEDIA_URL, p2_name)
    return file_url


def payment_record2_row_list(payment_record):
    datetime_format = '%Y/%m/%d-%H:%M:%S'

    return [
        payment_record.payer.get_full_name(),
        # payment_record.order,
        # payment_record.payment_type,
        payment_record.amount,
        payment_record.successful,
        payment_record.charged,
        payment_record.transaction_code,
        # payment_record.tref,
        # payment_record.payment_order_guid,
        payment_record.payment_order_id,
        # payment_record.payment_link,
        payment_record.called_back,
        payment_record.log_text,
        payment_record.created_at.strftime(datetime_format) if payment_record.created_at else None,
        payment_record.updated_at.strftime(datetime_format) if payment_record.updated_at else None,

    ]


def payment_record2():
    return payment_record2_row_list, [
            'پرداخت کننده',
            # 'order',
            # 'payment_type',
            'مقدار',
            'موفق',
            'شارژ',
            'کد تراکنش',
            # 'tref',
            # 'payment_order_guid',
            'شماره درخواست درگاه',
            # 'payment_link',
            'برگشت',
            'لاگ',
            'ساخته شده در',
            'ویرایش شده در'
        ], 'Payment2'


def payment_record_row_list(payment_record):
    datetime_format = '%Y/%m/%d-%H:%M:%S'

    return [
        payment_record.payer.get_full_name(),
        # payment_record.order,
        # payment_record.payment_type,
        payment_record.amount,
        payment_record.successful,
        payment_record.charged,
        payment_record.transaction_code,
        payment_record.tref,
        # payment_record.payment_order_guid,
        payment_record.payment_order_id,
        # payment_record.payment_link,
        payment_record.called_back,
        payment_record.log_text,
        payment_record.created_at.strftime(datetime_format) if payment_record.created_at else None,
        payment_record.updated_at.strftime(datetime_format) if payment_record.updated_at else None,
    ]


def payment_record():
    return payment_record_row_list, [
            'پرداخت کننده',
            # 'order',
            # 'payment_type',
            'مقدار',
            'موفق',
            'شارژ',
            'کد تراکنش',
            'کد مرجع بانک',
            # 'payment_order_guid',
            'شماره درخواست درگاه',
            # 'payment_link',
            'برگشت',
            'لاگ',
            'ساخته شده در',
            'ویرایش شده در'
        ], 'PaymentRecord'


def grant_request_row_list(grant_request):
    datetime_format = '%Y/%m/%d-%H:%M:%S'
    return [
        grant_request.sender.get_full_name(),
        grant_request.receiver.get_full_name(),
        grant_request.requested_amount if grant_request.requested_amount else None,
        grant_request.approved_amount if grant_request.approved_amount else None,
        grant_request.approved_datetime.strftime(datetime_format) if grant_request.approved_datetime else None,
        grant_request.datetime.strftime(datetime_format) if grant_request.datetime else None,
        grant_request.expiration_date.strftime(datetime_format) if grant_request.expiration_date else None,
        # grant_request.transaction.amount,
        grant_request.status,
    ]


def grant_request():
    return grant_request_row_list, [
            'فرستنده',
            'گیرنده',
            'مبلغ درخواست',
            'مبلغ تایید شده',
            'تاریخ تایید',
            'تاریخ ثبت درخواست',
            'تاریخ انقضا',
            # 'transaction',
            'وضعیت',
        ], 'GrantRequest'


def request():
    return request_row_list, [
            'شماره درخواست',
            'کاربر',
            'آزمایش',
            'پارامتر ها',
            'قیمت',
            'فوری',
            'تاریخ نتیجه',
            'توضیحات',
            # 'subject',
            'تکمیل شده',
            'ساخته شده در',
            # 'updated_at',
        ], 'Request'

def request_row_list(request):
    datetime_format = '%Y/%m/%d-%H:%M:%S'
    return [
        request.request_number if request.request_number else None,
        request.owner.get_full_name(),
        request.experiment.name,
        str([parameter.name for parameter in request.parameter.all()]) if request.parameter else None,
        request.price if request.price else None,
        request.is_urgent,
        request.delivery_date.strftime(datetime_format) if request.delivery_date else None,
        request.description if request.description else None,
        # request.subject if request.subject else None,
        request.is_completed,
        request.created_at.strftime(datetime_format) if request.created_at else None,
        # request.updated_at.strftime(datetime_format) if request.updated_at else None,
    ]

def user():
    return user_row_list, [
            'نام کاربری',
            'ایمیل',
            'کدملی',
            'نام',
            'نام خانوادگی',
            'نوع',
            'اکانت',
            'سمت',
            # 'دسترسی',
            'اعتبار',
        ], 'User'

def user_row_list(user):
    datetime_format = '%Y/%m/%d-%H:%M:%S'
    return [
        user.username.replace('+98','0'),
        user.email,
        user.national_id,
        user.first_name,
        user.last_name,
        user.user_type,
        user.account_type,
        str([role.name for role in user.role.all()]),
        # str([access_level.name for access_level in user.get_access_levels()]),
        user.balance,
    ]